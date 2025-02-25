import csv
import pandas as pd
import openpyxl
import firebase_admin
from firebase_admin import db
from datetime import datetime

cred_obj = firebase_admin.credentials.Certificate('safetyhelmet-6bdea-firebase-adminsdk.json')
default_app = firebase_admin.initialize_app(cred_obj, {
        'databaseURL': 'https://safetyhelmet-6bdea-default-rtdb.firebaseio.com/'})
refStart = db.reference("/AreaRisk")
area = refStart.child("Area").get()
time = refStart.child("Time").get()
areaSet = "Shah"
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("dataset.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active
string = "null"
count = 0
# Iterate the loop to read the cell values
for hours in range(0,24):
    for row in range(1, dataframe1.max_row):
        if row <= (dataframe1.max_row - 1):
            times = [dataframe1.cell(row=row, column=3).value]
            hour = str(times[0])
            result = int(hour[:2])
            # print(result)
            if hours == result:
                count = count + 1
    if count < 30:
        string = "LOW"
    elif count >= 30 and count < 50:
        string = "MEDIUM"
    elif count >= 50:
        string = "HIGH"
    print("Hour {0} : {1} -> {2}".format(hours, string,count))
    count = 0
    string = "null"




#     print(row,end='',sep='\t')
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value,end='')
#     print('')

# read by default 1st sheet of an excel file
# dataframe1 = pd.read_excel('2 WHEELERS  theft1.xlsx')
#
# print(dataframe1)

# with open('employee_birthday.txt', mode='r') as csv_file:
#     csv_reader = csv.DictReader(csv_file)
#     line_count = 0
#     for row in csv_reader:
#         if line_count == 0:
#             print(f'Column names are {", ".join(row)}')
#             line_count += 1
#         print(f'\t{row["name"]} works in the {row["department"]} department, and was born in {row["birthday month"]}.')
#         line_count += 1
#     print(f'Processed {line_count} lines.')
list_numbers = []
xls_file_list = os.listdir('xls_files')
for xls_file in xls_file_list:
    path = 'xls_files' + '/' + xls_file
    workbook = load_workbook(path)
    sheets = workbook.sheetnames
    for sheet in sheets:
        data = ''
        current_sheet = workbook[sheet]
        for i in range(1, current_sheet.max_row + 1):
            for j in range(1, current_sheet.max_column + 1):
                cell_obj = current_sheet.cell(row=i, column=j)
                data = data + ';' + str(cell_obj.value)

        pattern = '(\d{9})'
        for match in re.finditer(pattern, data):
            x = match.group()
            list_numbers.append(x)
list_numbers = list(set(list_numbers))
print(list_numbers)
# Using the read_only method since you're not gonna be editing the spreadsheet
workbook = load_workbook(filename="sample.xlsx", read_only=True)
sheet = workbook.active

products = []
reviews = []

# Using the values_only because you just want to return the cell value
for row in sheet.iter_rows(min_row=2, values_only=True):
    product = Product(id=row[PRODUCT_ID],
                      parent=row[PRODUCT_PARENT],
                      title=row[PRODUCT_TITLE],
                      category=row[PRODUCT_CATEGORY])
    products.append(product)

    # You need to parse the date from the spreadsheet into a datetime format
    spread_date = row[REVIEW_DATE]
    parsed_date = datetime.strptime(spread_date, "%Y-%m-%d")

    review = Review(id=row[REVIEW_ID],
                    customer_id=row[REVIEW_CUSTOMER],
                    stars=row[REVIEW_STARS],
                    headline=row[REVIEW_HEADLINE],
                    body=row[REVIEW_BODY],
                    date=parsed_date)
    reviews.append(review)

print(products[0])
print(reviews[0])